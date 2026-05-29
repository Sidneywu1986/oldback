from typing import Optional
from datetime import datetime
from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from fastapi.responses import FileResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.core.database import get_db
from app.core.dependencies import get_current_user
from app.models.user import User
from app.services.report_service import ReportService
from app.schemas.base import BaseResponse

router = APIRouter(prefix="/api/v1", tags=["报表管理"])


@router.get("/reports/recycle-trend", summary="获取回收趋势报表")
def get_recycle_trend(
    days: int = Query(30, ge=7, le=365, description="查询天数"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    data = ReportService.get_recycle_trend(db, days)
    return BaseResponse(data=data)


@router.get("/reports/master-performance", summary="获取师傅绩效报表")
def get_master_performance(
    start_time: Optional[datetime] = Query(None, description="开始时间"),
    end_time: Optional[datetime] = Query(None, description="结束时间"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    data = ReportService.get_master_performance(db, start_time, end_time)
    return BaseResponse(data=data)


@router.get("/reports/parts-category", summary="获取旧件分类统计")
def get_parts_category_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    data = ReportService.get_parts_category_stats(db)
    return BaseResponse(data=data)


@router.get("/reports/fund", summary="获取资金报表")
def get_fund_report(
    start_time: Optional[datetime] = Query(None, description="开始时间"),
    end_time: Optional[datetime] = Query(None, description="结束时间"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    data = ReportService.get_fund_report(db, start_time, end_time)
    return BaseResponse(data=data)


@router.get("/reports/points", summary="获取积分报表")
def get_points_report(
    start_time: Optional[datetime] = Query(None, description="开始时间"),
    end_time: Optional[datetime] = Query(None, description="结束时间"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    data = ReportService.get_points_report(db, start_time, end_time)
    return BaseResponse(data=data)


@router.get("/reports/monthly", summary="获取月度综合报表")
def get_monthly_report(
    year: int = Query(..., ge=2020, description="年份"),
    month: int = Query(..., ge=1, le=12, description="月份"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    data = ReportService.get_monthly_report(db, year, month)
    return BaseResponse(data=data)


@router.get("/reports/master-performance/export", summary="导出师傅绩效报表")
def export_master_performance(
    start_time: Optional[datetime] = Query(None, description="开始时间"),
    end_time: Optional[datetime] = Query(None, description="结束时间"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    data = ReportService.get_master_performance(db, start_time, end_time)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "师傅绩效报表"
    
    headers = ["师傅ID", "姓名", "手机号", "等级", "信用分", "回收次数", "累计积分", "累计金额"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    for row_num, record in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=record["master_id"])
        ws.cell(row=row_num, column=2, value=record["name"])
        ws.cell(row=row_num, column=3, value=record["phone"])
        ws.cell(row=row_num, column=4, value=record["level"])
        ws.cell(row=row_num, column=5, value=record["credit_score"])
        ws.cell(row=row_num, column=6, value=record["recycle_count"])
        ws.cell(row=row_num, column=7, value=float(record["total_points"]))
        ws.cell(row=row_num, column=8, value=float(record["total_amount"]))
    
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return FileResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"师傅绩效报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


@router.get("/reports/fund/export", summary="导出资金报表")
def export_fund_report(
    start_time: Optional[datetime] = Query(None, description="开始时间"),
    end_time: Optional[datetime] = Query(None, description="结束时间"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    data = ReportService.get_fund_report(db, start_time, end_time)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "资金报表"
    
    ws.cell(row=1, column=1, value="资金报表概览")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    summary_data = [
        ("总收入", data["total_income"]),
        ("总提现", data["total_withdraw"]),
        ("总调整", data["total_adjust"]),
        ("余额变动", data["balance_change"]),
    ]
    
    for row_num, (label, value) in enumerate(summary_data, 3):
        ws.cell(row=row_num, column=1, value=label)
        ws.cell(row=row_num, column=2, value=float(value))
    
    ws.cell(row=6, column=1, value="交易类型统计")
    ws.cell(row=6, column=1).font = Font(bold=True)
    
    txn_headers = ["交易类型", "交易次数", "交易金额"]
    for col_num, header in enumerate(txn_headers, 1):
        cell = ws.cell(row=7, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    for row_num, txn in enumerate(data["txn_by_type"], 8):
        ws.cell(row=row_num, column=1, value=txn["txn_type_label"])
        ws.cell(row=row_num, column=2, value=txn["count"])
        ws.cell(row=row_num, column=3, value=float(txn["total_amount"]))
    
    for col_num in range(1, 4):
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return FileResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"资金报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


@router.get("/reports/monthly/export", summary="导出月度综合报表")
def export_monthly_report(
    year: int = Query(..., ge=2020, description="年份"),
    month: int = Query(..., ge=1, le=12, description="月份"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    data = ReportService.get_monthly_report(db, year, month)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{data['period']}综合报表"
    
    ws.cell(row=1, column=1, value=f"{data['period']}综合报表")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    ws.cell(row=3, column=1, value="回收统计")
    ws.cell(row=3, column=1).font = Font(bold=True)
    
    recycle_data = [
        ("订单总数", data["recycle"]["total_orders"]),
        ("积分总额", data["recycle"]["total_points"]),
        ("回收金额", data["recycle"]["total_amount"]),
    ]
    
    for row_num, (label, value) in enumerate(recycle_data, 4):
        ws.cell(row=row_num, column=1, value=label)
        ws.cell(row=row_num, column=2, value=value if isinstance(value, int) else float(value))
    
    ws.cell(row=7, column=1, value="师傅统计")
    ws.cell(row=7, column=1).font = Font(bold=True)
    ws.cell(row=8, column=1, value="活跃师傅数")
    ws.cell(row=8, column=2, value=data["masters"]["active_masters"])
    
    ws.cell(row=10, column=1, value="资金统计")
    ws.cell(row=10, column=1).font = Font(bold=True)
    
    fund_data = [
        ("奖励发放", data["fund"]["total_award"]),
        ("提现金额", data["fund"]["total_withdraw"]),
    ]
    
    for row_num, (label, value) in enumerate(fund_data, 11):
        ws.cell(row=row_num, column=1, value=label)
        ws.cell(row=row_num, column=2, value=float(value))
    
    for col_num in range(1, 3):
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return FileResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"月度综合报表_{data['year']}{str(data['month']).zfill(2)}.xlsx"
    )